from read import read_files
from database import query_snowflake
from config import *
from datetime import datetime
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# add action to action column -- contact sales, shut it down, etc
def add_action(row):
    if row['TRIAL_STATUS'] == "Expired" and row['ORDER_ID'].startswith('4E'):
        if row['IN_CX_NETWORK'] == "Yes":
            if row['IS_EA_ORG'] == "Yes":
                return 'Contact Sales'
            elif "MX" in row['PROD_CODE'] or "MS" in row['PROD_CODE']:
                if row['SUPPORT_SCORE'] in ["0", "1"]:
                    return "Contact Sales"
                else: 
                    return "Remove node & shut down"
            else: 
                return "Remove node & shut down"
        elif row['IN_CX_NETWORK'] == "ShutDown":
            return "None"
        else: 
            return "Shut it down"
    else:
        return "DO NOT SHUTDOWN"
        

# change purchased trial's status to Purchased 
def update_trial_status(row):
    if not row['ORDER_ID'].startswith('4E'):
        row['TRIAL_STATUS'] = "Purchased"
    return row


def calculate_optimal_width(value, font=Font(name='Calibri', size=11)):
    # calculate the optimal width for a column in Excel based on the content
    if value is None:
        return 0
    # calculate the length of the string
    length = len(str(value))
    # approximate character width based on font and size
    character_width = 1  # Adjust this based on your font and size
    # calculate the optimal width
    optimal_width = (length + 2) * character_width
    return optimal_width


def apply_conditional_formatting(sheet, df):
    # Adjust column width based on the widest content (header or data) in each column
    for col in sheet.columns:
        max_width = 0
        column_letter = col[0].column_letter  # Get the column letter
        # Calculate width based on column header
        header_width = calculate_optimal_width(col[0].value)
        max_width = max(max_width, header_width)
        # Calculate width based on data rows
        for cell in col[1:]:
            cell_width = calculate_optimal_width(cell.value)
            max_width = max(max_width, cell_width)
        # Set the column width
        sheet.column_dimensions[column_letter].width = max_width



    if 'PROD_CODE' in df.columns:
        prod_code_col_idx = df.columns.get_loc('PROD_CODE') + 1
        for row in sheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=prod_code_col_idx, max_col=prod_code_col_idx):
            prod_code_value = row[0].value
            if prod_code_value and ('MX' in str(prod_code_value) or 'MS' in str(prod_code_value)):
                for cell in row:
                    cell.fill = PatternFill(start_color="B2CBDE", end_color="B2CBDE", fill_type="solid")
    else:
        print("PROD_CODE column not found in DataFrame.")
    

    if 'IS_EA_ORG' in df.columns:
        ea_org_col_idx = df.columns.get_loc('IS_EA_ORG') + 1
        for row in sheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=ea_org_col_idx, max_col=ea_org_col_idx):
            ea_org_value = row[0].value
            if ea_org_value and "Yes" in str(ea_org_value):
                for cell in row:
                    cell.fill = PatternFill(start_color="FDFD96", end_color="FDFD96", fill_type="solid")
    else:
        print("IS_EA_ORG column not found in DataFrame.")


    if 'SUPPORT_SCORE' in df.columns:
        sup_score_col_idx = df.columns.get_loc('SUPPORT_SCORE') + 1
        for row in sheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=sup_score_col_idx, max_col=sup_score_col_idx):
            sup_score_value = row[0].value
            if sup_score_value and ('0' in str(sup_score_value) or '1' in str(sup_score_value)):
                for cell in row:
                    cell.fill = PatternFill(start_color="FFD1DC", end_color="FFD1DC", fill_type="solid")
    else:
        print("SUPPORT_SCORE column not found in DataFrame.")


    if 'ACTION' in df.columns: 
        action_col_idx = df.columns.get_loc('ACTION') + 1
        for row in sheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=action_col_idx, max_col=action_col_idx):
            action_value = row[0].value
            if action_value and "DO NOT SHUTDOWN" in str(action_value):
                for cell in row:
                    cell.fill = PatternFill(start_color="FF483F", end_color="FF483F", fill_type="solid")
            if action_value and "Contact Sales" in str(action_value):
                for cell in row:
                    cell.fill = PatternFill(start_color="FFC17A", end_color="FFC17A", fill_type="solid")
    else:
        print("ACTION column not found in DataFrame.")


    if 'TRIAL_STATUS' in df.columns: 
        status_col_idx = df.columns.get_loc('TRIAL_STATUS') + 1
        for row in sheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=status_col_idx, max_col=status_col_idx):
            status_value = row[0].value
            if status_value and ("Purchased" in str(status_value)) or ("Not Expired" in str(status_value)):
                for cell in row:
                    cell.fill = PatternFill(start_color="A7CF97", end_color="A7CF97", fill_type="solid")              
    else:
        print("TRIAL_STATUS column not found in DataFrame.")




def main():
    order_data = read_files(SOURCE_DIRECTORY)
    if not order_data:
        print("No order IDs found, exiting.")
        return
    
    for file, order_id_list in order_data.items():
        result_df = query_snowflake(order_id_list)
        # change null values to "NA"
        result_df = result_df.fillna(value="NA")
        print("Dataframe created")

        if result_df.empty:
            print("No data returned from Snowflake for file {file}, skipping.")
            continue

        # create list of serials from original excel file
        original_serials = pd.read_excel(file)['SERIAL'].to_list()
        
        # remove serials that were not on original excel file (so there aren't extra devices from a particular order)
        result_df = result_df[result_df['SERIAL'].isin(original_serials)]
        
        # drop any duplicate serials (if the query pulls a 5S number for puchased trial)
        result_df = result_df.drop_duplicates(subset=['SERIAL'])

        # sort DataFrame by 'ORDER_ID' so order numbers are grouped together
        sorted_df = result_df.sort_values(by='ORDER_ID')

        # change trial status from expired to purchased if 5S or 4S order number
        sorted_df = sorted_df.apply(update_trial_status, axis=1)

        # add action column to excel sheet to show next steps for each device
        sorted_df['ACTION'] = sorted_df.apply(add_action, axis=1)


        base_filename = os.path.basename(file)
        output_filename = f"updatedv2_{base_filename}"

        output_path = os.path.join(TARGET_DIRECTORY, output_filename)

        # write DataFrame to excel
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        sorted_df.to_excel(writer, index=False)


        # apply conditional formatting
        workbook = writer.book
        sheet_name = "Sheet1"
        sheet = workbook[sheet_name]
        apply_conditional_formatting(sheet, sorted_df)

        # save the workbook after applying conditional formatting
        workbook.save(output_path)

        print(f"Export successful: {output_path}")



if __name__ == "__main__":
    main()